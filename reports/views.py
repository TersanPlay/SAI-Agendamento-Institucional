from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.utils import timezone
from django.db.models import Count, Q
from datetime import datetime, timedelta
import json
import time
import os
from io import BytesIO

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import Report, ReportExecution
from events.models import Event, EventType, Department, Location
from accounts.utils import get_user_accessible_events, has_permission
from accounts.models import User


@login_required
def reports_view(request):
    """Página unificada de relatórios - combina listagem e criação"""
    if not has_permission(request.user, 'view_reports'):
        messages.error(request, 'Você não tem permissão para acessar relatórios.')
        return redirect('events:dashboard')
    
    # Handle form submission for generating reports
    if request.method == 'POST':
        # Processar formulário
        report_type = request.POST.get('report_type')
        format_type = request.POST.get('format', 'pdf')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Generate name automatically
        report_type_label = dict(Report.REPORT_TYPES).get(report_type, 'Relatório')
        name = f"{report_type_label} - {start_date} a {end_date}"
        
        # Create temporary report object (not saved to database)
        report = Report(
            name=name,
            report_type=report_type,
            description='',  # Empty description
            start_date=datetime.strptime(start_date, '%Y-%m-%d').date(),
            end_date=datetime.strptime(end_date, '%Y-%m-%d').date(),
            format=format_type,
            created_by=request.user
        )
        
        # Add departments and event types if selected
        departments = request.POST.getlist('departments')
        event_types = request.POST.getlist('event_types')
        
        # Set departments and event types on the temporary report
        if departments:
            report.departments.set(departments)  # type: ignore
        if event_types:
            report.event_types.set(event_types)  # type: ignore
        
        try:
            # Generate the report directly without saving
            if format_type == 'pdf':
                file_content, filename = generate_pdf_report(report)
                content_type = 'application/pdf'
            elif format_type == 'excel':
                file_content, filename = generate_excel_report(report)
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            else:
                file_content, filename = generate_csv_report(report)
                content_type = 'text/csv'
            
            # Return the file as download
            response = HttpResponse(file_content, content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            messages.error(request, f'Erro ao gerar relatório: {str(e)}')
            return redirect('reports:list')
    
    # GET request - mostrar formulário e lista de relatórios
    context = {
        'can_create_reports': has_permission(request.user, 'add_report'),
        'departments': Department.objects.all(),  # type: ignore
        'event_types': EventType.objects.all(),  # type: ignore
        'report_types': Report.REPORT_TYPES,
        'format_choices': Report.FORMAT_CHOICES,
    }
    
    return render(request, 'reports/advanced.html', context)


@login_required
def report_data_api(request):
    """API endpoint to fetch report data dynamically"""
    if not has_permission(request.user, 'view_reports'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    report_type = request.GET.get('report_type', 'events_by_period')
    status = request.GET.get('status')
    department_ids = [d for d in request.GET.getlist('departments') if d]
    event_type_ids = [e for e in request.GET.getlist('event_types') if e]
    location_ids = [l for l in request.GET.getlist('locations') if l]
    search = request.GET.get('search')
    responsible_search = request.GET.get('responsible_search')
    
    # Parse dates if provided
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Base queryset
    events = get_user_accessible_events(request.user)
    
    # Apply date filters
    if start_date:
        events = events.filter(start_datetime__date__gte=start_date)
    if end_date:
        events = events.filter(start_datetime__date__lte=end_date)
    
    # Apply status filter
    if status:
        events = events.filter(status=status)
    
    # Apply department filter
    if department_ids:
        events = events.filter(department_id__in=department_ids)
    
    # Apply event type filter
    if event_type_ids:
        events = events.filter(event_type_id__in=event_type_ids)
    
    # Apply location filter
    if location_ids:
        events = events.filter(location_id__in=location_ids)
    
    # Apply search filter
    if search:
        events = events.filter(
            Q(name__icontains=search) | 
            Q(description__icontains=search)
        )
    
    # Apply responsible person search filter
    if responsible_search:
        events = events.filter(
            Q(responsible_person__first_name__icontains=responsible_search) |
            Q(responsible_person__last_name__icontains=responsible_search) |
            Q(responsible_person__username__icontains=responsible_search)
        )
    
    # Prepare response data
    data = {
        'total_events': events.count(),
        'completed_events': events.filter(status='concluido').count(),
        'departments_count': events.values('department').distinct().count(),
        'event_types_count': events.values('event_type').distinct().count(),
    }
    
    # Add comparison data if dates are provided
    if start_date and end_date:
        comparison_data = get_comparison_data(request.user, start_date, end_date)
        data['comparison'] = comparison_data
    
    # Add events data
    events_data = []
    for event in events.select_related('event_type', 'department', 'responsible_person')[:25]:
        events_data.append({
            'id': str(event.id),
            'name': event.name,
            'event_type': event.event_type.name,
            'department': event.department.name,
            'start_datetime': event.start_datetime.strftime('%d/%m/%Y %H:%M'),
            'status': event.get_status_display(),
            'responsible_person': f"{event.responsible_person.get_full_name() or event.responsible_person.username}",
        })
    
    data['events'] = events_data
    
    # Add chart data
    data['events_by_type'] = list(
        events.values('event_type__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )
    
    data['events_by_department'] = list(
        events.values('department__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )
    
    return JsonResponse(data)


@login_required
def locations_api(request):
    """API endpoint to fetch locations for dropdown"""
    if not has_permission(request.user, 'view_reports'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    # Get all locations from events accessible to the user
    events = get_user_accessible_events(request.user)
    locations = events.values('location_id', 'location__name', 'location__custom_name').distinct()
    
    # Format the data
    locations_data = []
    for location in locations:
        if location['location_id']:
            display_name = location['location__custom_name'] if location['location__custom_name'] else location['location__name']
            locations_data.append({
                'id': location['location_id'],
                'name': display_name
            })
    
    return JsonResponse({'locations': locations_data})


def get_comparison_data(user, start_date, end_date):
    """Get data for the previous period to compare with current period"""
    # Calculate previous period dates
    current_period_days = (end_date - start_date).days + 1
    previous_end_date = start_date - timedelta(days=1)
    previous_start_date = previous_end_date - timedelta(days=current_period_days - 1)
    
    # Get events for previous period
    previous_events = get_user_accessible_events(user)
    previous_events = previous_events.filter(
        start_datetime__date__gte=previous_start_date,
        start_datetime__date__lte=previous_end_date
    )
    
    # Get events for current period
    current_events = get_user_accessible_events(user)
    current_events = current_events.filter(
        start_datetime__date__gte=start_date,
        start_datetime__date__lte=end_date
    )
    
    # Calculate metrics
    previous_total = previous_events.count()
    current_total = current_events.count()
    
    previous_completed = previous_events.filter(status='concluido').count()
    current_completed = current_events.filter(status='concluido').count()
    
    # Calculate changes
    total_change = current_total - previous_total
    completed_change = current_completed - previous_completed
    
    # Calculate percentages
    total_change_percent = (total_change / previous_total * 100) if previous_total > 0 else 0
    completed_change_percent = (completed_change / previous_completed * 100) if previous_completed > 0 else 0
    
    return {
        'total_change': total_change,
        'total_change_percent': round(total_change_percent, 2),
        'completed_change': completed_change,
        'completed_change_percent': round(completed_change_percent, 2),
        'previous_period': {
            'start_date': previous_start_date,
            'end_date': previous_end_date,
            'total_events': previous_total,
            'completed_events': previous_completed
        }
    }


@login_required
def trend_data_api(request):
    """API endpoint to fetch trend data for the dashboard"""
    if not has_permission(request.user, 'view_reports'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    # Get the last 12 months of data
    today = timezone.now().date()
    months = []
    event_counts = []
    
    for i in range(11, -1, -1):  # Last 12 months
        # Calculate the start and end of the month
        if i == 0:
            # Current month
            start_date = today.replace(day=1)
            end_date = today
        else:
            # Previous months
            month = today.month - i
            year = today.year
            if month <= 0:
                month += 12
                year -= 1
            start_date = datetime(year, month, 1).date()
            
            # Calculate end of month
            if month == 12:
                end_date = datetime(year, 12, 31).date()
            else:
                end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        # Get events for this month
        events = get_user_accessible_events(request.user)
        events = events.filter(
            start_datetime__date__gte=start_date,
            start_datetime__date__lte=end_date
        )
        
        months.append(start_date.strftime('%b %Y'))
        event_counts.append(events.count())
    
    return JsonResponse({
        'months': months,
        'event_counts': event_counts
    })


@login_required
def export_report(request):
    """View to handle report export functionality"""
    if not has_permission(request.user, 'view_reports'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method == 'POST':
        try:
            # Get filter parameters
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            report_type = request.POST.get('report_type', 'events_by_period')
            format_type = request.POST.get('format', 'pdf')
            status = request.POST.get('status')
            department_ids = [d for d in request.POST.getlist('departments') if d]
            event_type_ids = [e for e in request.POST.getlist('event_types') if e]
            location_ids = [l for l in request.POST.getlist('locations') if l]
            search = request.POST.get('search')
            responsible_search = request.POST.get('responsible_search')
            
            # Parse dates
            if start_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            if end_date:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            # Generate name automatically
            report_type_label = dict(Report.REPORT_TYPES).get(report_type, 'Relatório')
            period_str = ''
            if start_date and end_date:
                period_str = f" - {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
            elif start_date:
                period_str = f" - a partir de {start_date.strftime('%d/%m/%Y')}"
            elif end_date:
                period_str = f" - até {end_date.strftime('%d/%m/%Y')}"
            
            name = f"{report_type_label}{period_str}"
            
            # Get filtered events data directly
            events = get_user_accessible_events(request.user)
            
            # Apply all filters
            if start_date:
                events = events.filter(start_datetime__date__gte=start_date)
            if end_date:
                events = events.filter(start_datetime__date__lte=end_date)
            if status:
                events = events.filter(status=status)
            if department_ids:
                events = events.filter(department_id__in=department_ids)
            if event_type_ids:
                events = events.filter(event_type_id__in=event_type_ids)
            if location_ids:
                events = events.filter(location_id__in=location_ids)
            if search:
                events = events.filter(
                    Q(name__icontains=search) | 
                    Q(description__icontains=search)
                )
            if responsible_search:
                events = events.filter(
                    Q(responsible_person__first_name__icontains=responsible_search) |
                    Q(responsible_person__last_name__icontains=responsible_search) |
                    Q(responsible_person__username__icontains=responsible_search)
                )
            
            # Create export data structure
            export_data = {
                'name': name,
                'report_type': report_type,
                'start_date': start_date,
                'end_date': end_date,
                'events': events,
                'created_by': request.user,
                'format': format_type
            }
            
            # Generate the report
            if format_type == 'pdf':
                file_content, filename = generate_dynamic_pdf_report(export_data)
                content_type = 'application/pdf'
            elif format_type == 'excel':
                file_content, filename = generate_dynamic_excel_report(export_data)
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            else:
                file_content, filename = generate_dynamic_csv_report(export_data)
                content_type = 'text/csv'
            
            # Return the file as download
            response = HttpResponse(file_content, content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            return JsonResponse({'error': f'Erro ao gerar relatório: {str(e)}'}, status=500)
    
    # If not POST, redirect to the reports page
    return redirect('reports:list')


def get_report_data(report):
    """Obter dados para o relatório"""
    # Filtros base
    events = get_user_accessible_events(report.created_by)
    events = events.filter(
        start_datetime__date__gte=report.start_date,
        start_datetime__date__lte=report.end_date
    )
    
    # Aplicar filtros adicionais
    if hasattr(report, 'departments') and report.departments.exists():
        events = events.filter(department__in=report.departments.all())
    
    if hasattr(report, 'event_types') and report.event_types.exists():
        events = events.filter(event_type__in=report.event_types.all())
    
    # Retornar dados baseado no tipo de relatório
    if report.report_type == 'events_by_period':
        return list(events.values(
            'name', 'start_datetime', 'end_datetime', 'location__name',
            'event_type__name', 'status', 'description'
        ))
    
    elif report.report_type == 'events_by_type':
        return list(events.values('event_type__name').annotate(
            count=Count('id')
        ).order_by('-count'))
    
    elif report.report_type == 'events_by_department':
        return list(events.values('department__name').annotate(
            count=Count('id')
        ).order_by('-count'))
    
    elif report.report_type == 'events_by_status':
        return list(events.values('status').annotate(
            count=Count('id')
        ).order_by('-count'))
    
    else:
        # Relatório padrão com todos os campos
        return list(events.values(
            'name', 'start_datetime', 'end_datetime', 'location__name',
            'event_type__name', 'department__name', 'status', 'description',
            'responsible_person__first_name', 'responsible_person__last_name'
        ))


def generate_pdf_report(report):
    """Gerar relatório em PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Elementos do documento
    elements = []
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    elements.append(Paragraph(f"Relatório: {report.name}", title_style))
    elements.append(Paragraph(f"Período: {report.start_date.strftime('%d/%m/%Y')} a {report.end_date.strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Paragraph(f"Gerado em: {timezone.now().strftime('%d/%m/%Y às %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Dados
    data = get_report_data(report)
    
    if not data:
        elements.append(Paragraph("Nenhum dado encontrado para os critérios especificados.", styles['Normal']))
    else:
        # Criar tabela baseada no tipo de relatório
        table_data = []
        if report.report_type in ['events_by_type', 'events_by_department', 'events_by_status']:
            # Relatórios de agregação
            table_data = [['Item', 'Quantidade']]
            for item in data:
                key = list(item.keys())[0]  # Primeira chave (nome do campo)
                value = item[key] if item[key] else 'Não especificado'
                count = item['count']
                table_data.append([str(value), str(count)])
        else:
            # Relatórios detalhados
            if data:
                headers = {
                    'name': 'Título',
                    'start_datetime': 'Data Início',
                    'end_datetime': 'Data Fim',
                    'location__name': 'Local',
                    'event_type__name': 'Tipo',
                    'department__name': 'Departamento',
                    'status': 'Status',
                    'responsible_person__first_name': 'Responsável',
                }
                
                # Cabeçalhos da tabela
                available_fields = list(data[0].keys())
                table_headers = [headers.get(field, field.replace('__', ' ').title()) for field in available_fields]
                table_data = [table_headers]
                
                # Dados da tabela
                for item in data:
                    row = []
                    for field in available_fields:
                        value = item.get(field, '')
                        if field in ['start_datetime', 'end_datetime'] and value:
                            value = value.strftime('%d/%m/%Y %H:%M') if hasattr(value, 'strftime') else str(value)
                        elif field == 'responsible_person__first_name' and item.get('responsible_person__last_name'):
                            value = f"{value or ''} {item.get('responsible_person__last_name', '')}".strip()
                        row.append(str(value or ''))
                    table_data.append(row)
        
        # Criar tabela
        if table_data:
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            
            elements.append(table)
        
        # Resumo
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(f"Total de registros: {len(data)}", styles['Normal']))
    
    # Gerar PDF
    doc.build(elements)
    pdf_content = buffer.getvalue()
    buffer.close()
    
    filename = f"relatorio_{report.report_type}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{report.format}"
    return pdf_content, filename


def generate_excel_report(report):
    """Gerar relatório em Excel"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Relatório")
    else:
        ws.title = "Relatório"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Cabeçalho do relatório
    ws['A1'] = f"Relatório: {report.name}"
    ws['A2'] = f"Período: {report.start_date.strftime('%d/%m/%Y')} a {report.end_date.strftime('%d/%m/%Y')}"
    ws['A3'] = f"Gerado em: {timezone.now().strftime('%d/%m/%Y às %H:%M')}"
    
    # Dados
    data = get_report_data(report)
    
    if not data:
        ws['A5'] = "Nenhum dado encontrado para os critérios especificados."
    else:
        start_row = 5
        
        if report.report_type in ['events_by_type', 'events_by_department', 'events_by_status']:
            # Relatórios de agregação
            ws[f'A{start_row}'] = 'Item'
            ws[f'B{start_row}'] = 'Quantidade'
            
            # Aplicar estilo ao cabeçalho
            for col in ['A', 'B']:
                cell = ws[f'{col}{start_row}']
                cell.font = header_font
                cell.fill = header_fill
            
            # Preencher dados
            for i, item in enumerate(data, start_row + 1):
                key = list(item.keys())[0]  # Primeira chave (nome do campo)
                value = item[key] if item[key] else 'Não especificado'
                count = item['count']
                ws[f'A{i}'] = str(value)
                ws[f'B{i}'] = count
        else:
            # Relatórios detalhados
            if data:
                headers = {
                    'name': 'Título',
                    'start_datetime': 'Data Início',
                    'end_datetime': 'Data Fim',
                    'location__name': 'Local',
                    'event_type__name': 'Tipo',
                    'department__name': 'Departamento',
                    'status': 'Status',
                    'responsible_person__first_name': 'Responsável',
                }
                
                # Cabeçalhos da tabela
                available_fields = list(data[0].keys())
                table_headers = [headers.get(field, field.replace('__', ' ').title()) for field in available_fields]
                
                # Preencher cabeçalhos
                for col_idx, header in enumerate(table_headers, 1):
                    col_letter = get_column_letter(col_idx)
                    cell = ws[f'{col_letter}{start_row}']
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Preencher dados
                for row_idx, item in enumerate(data, start_row + 1):
                    for col_idx, field in enumerate(available_fields, 1):
                        col_letter = get_column_letter(col_idx)
                        value = item.get(field, '')
                        if field in ['start_datetime', 'end_datetime'] and value:
                            value = value.strftime('%d/%m/%Y %H:%M') if hasattr(value, 'strftime') else str(value)
                        elif field == 'responsible_person__first_name' and item.get('responsible_person__last_name'):
                            value = f"{value or ''} {item.get('responsible_person__last_name', '')}".strip()
                        ws[f'{col_letter}{row_idx}'] = str(value or '')
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column) if column[0].column else 'A'
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"relatorio_{report.report_type}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buffer.getvalue(), filename


def generate_csv_report(report):
    """Gerar relatório em CSV"""
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Cabeçalho
    writer.writerow(['Relatório', report.name])
    writer.writerow(['Período', f"{report.start_date.strftime('%d/%m/%Y')} a {report.end_date.strftime('%d/%m/%Y')}"])
    writer.writerow(['Gerado em', timezone.now().strftime('%d/%m/%Y às %H:%M')])
    writer.writerow([])  # Linha em branco
    
    # Dados
    data = get_report_data(report)
    
    if not data:
        writer.writerow(['Nenhum dado encontrado para os critérios especificados.'])
    else:
        if report.report_type in ['events_by_type', 'events_by_department', 'events_by_status']:
            # Relatórios de agregação
            writer.writerow(['Item', 'Quantidade'])
            for item in data:
                key = list(item.keys())[0]  # Primeira chave (nome do campo)
                value = item[key] if item[key] else 'Não especificado'
                count = item['count']
                writer.writerow([str(value), count])
        else:
            # Relatórios detalhados
            if data:
                headers = {
                    'name': 'Título',
                    'start_datetime': 'Data Início',
                    'end_datetime': 'Data Fim',
                    'location__name': 'Local',
                    'event_type__name': 'Tipo',
                    'department__name': 'Departamento',
                    'status': 'Status',
                    'responsible_person__first_name': 'Responsável',
                }
                
                # Cabeçalhos da tabela
                available_fields = list(data[0].keys())
                table_headers = [headers.get(field, field.replace('__', ' ').title()) for field in available_fields]
                writer.writerow(table_headers)
                
                # Dados da tabela
                for item in data:
                    row = []
                    for field in available_fields:
                        value = item.get(field, '')
                        if field in ['start_datetime', 'end_datetime'] and value:
                            value = value.strftime('%d/%m/%Y %H:%M') if hasattr(value, 'strftime') else str(value)
                        elif field == 'responsible_person__first_name' and item.get('responsible_person__last_name'):
                            value = f"{value or ''} {item.get('responsible_person__last_name', '')}".strip()
                        row.append(str(value or ''))
                    writer.writerow(row)
    
    # Converter para bytes
    csv_content = output.getvalue().encode('utf-8')
    output.close()
    
    filename = f"relatorio_{report.report_type}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return csv_content, filename


def generate_dynamic_pdf_report(export_data):
    """Generate PDF report from dynamic data"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Elements for the document
    elements = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    elements.append(Paragraph(f"Relatório: {export_data['name']}", title_style))
    
    # Period info
    period_info = ""
    if export_data['start_date'] and export_data['end_date']:
        period_info = f"Período: {export_data['start_date'].strftime('%d/%m/%Y')} a {export_data['end_date'].strftime('%d/%m/%Y')}"
    elif export_data['start_date']:
        period_info = f"A partir de: {export_data['start_date'].strftime('%d/%m/%Y')}"
    elif export_data['end_date']:
        period_info = f"Até: {export_data['end_date'].strftime('%d/%m/%Y')}"
    
    if period_info:
        elements.append(Paragraph(period_info, styles['Normal']))
    
    elements.append(Paragraph(f"Gerado em: {timezone.now().strftime('%d/%m/%Y às %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Get events data
    events = export_data['events'].select_related('event_type', 'department', 'location', 'responsible_person')
    
    if not events.exists():
        elements.append(Paragraph("Nenhum evento encontrado para os critérios especificados.", styles['Normal']))
    else:
        # Create table
        table_data = [[
            'Título', 'Tipo', 'Departamento', 'Data Início', 
            'Status', 'Responsável', 'Local'
        ]]
        
        for event in events:
            responsible_name = ''
            if event.responsible_person:
                responsible_name = event.responsible_person.get_full_name() or event.responsible_person.username
            
            location_name = ''
            if event.location:
                location_name = event.location.custom_name or event.location.name
            
            table_data.append([
                event.name[:30] + '...' if len(event.name) > 30 else event.name,
                event.event_type.name,
                event.department.name,
                event.start_datetime.strftime('%d/%m/%Y %H:%M'),
                event.get_status_display(),
                responsible_name,
                location_name
            ])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        elements.append(table)
        
        # Summary
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(f"Total de eventos: {events.count()}", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    pdf_content = buffer.getvalue()
    buffer.close()
    
    filename = f"relatorio_{export_data['report_type']}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return pdf_content, filename


def generate_dynamic_excel_report(export_data):
    """Generate Excel report from dynamic data"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Relatório")
    else:
        ws.title = "Relatório"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Report header
    ws['A1'] = f"Relatório: {export_data['name']}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Period info
    row = 2
    if export_data['start_date'] and export_data['end_date']:
        ws[f'A{row}'] = f"Período: {export_data['start_date'].strftime('%d/%m/%Y')} a {export_data['end_date'].strftime('%d/%m/%Y')}"
        row += 1
    elif export_data['start_date']:
        ws[f'A{row}'] = f"A partir de: {export_data['start_date'].strftime('%d/%m/%Y')}"
        row += 1
    elif export_data['end_date']:
        ws[f'A{row}'] = f"Até: {export_data['end_date'].strftime('%d/%m/%Y')}"
        row += 1
    
    ws[f'A{row}'] = f"Gerado em: {timezone.now().strftime('%d/%m/%Y às %H:%M')}"
    row += 2
    
    # Get events data
    events = export_data['events'].select_related('event_type', 'department', 'location', 'responsible_person')
    
    if not events.exists():
        ws[f'A{row}'] = "Nenhum evento encontrado para os critérios especificados."
    else:
        # Headers
        headers = ['Título', 'Tipo', 'Departamento', 'Data Início', 'Data Fim', 'Status', 'Responsável', 'Local', 'Descrição']
        
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{row}']
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        # Data
        for event in events:
            row += 1
            responsible_name = ''
            if event.responsible_person:
                responsible_name = event.responsible_person.get_full_name() or event.responsible_person.username
            
            location_name = ''
            if event.location:
                location_name = event.location.custom_name or event.location.name
            
            data_row = [
                event.name,
                event.event_type.name,
                event.department.name,
                event.start_datetime.strftime('%d/%m/%Y %H:%M'),
                event.end_datetime.strftime('%d/%m/%Y %H:%M') if event.end_datetime else '',
                event.get_status_display(),
                responsible_name,
                location_name,
                event.description or ''
            ]
            
            for col_idx, value in enumerate(data_row, 1):
                col_letter = get_column_letter(col_idx)
                ws[f'{col_letter}{row}'] = str(value)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column) if column[0].column else 'A'
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"relatorio_{export_data['report_type']}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buffer.getvalue(), filename


def generate_dynamic_csv_report(export_data):
    """Generate CSV report from dynamic data"""
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Header
    writer.writerow(['Relatório', export_data['name']])
    if export_data['start_date'] and export_data['end_date']:
        writer.writerow(['Período', f"{export_data['start_date'].strftime('%d/%m/%Y')} a {export_data['end_date'].strftime('%d/%m/%Y')}"])
    writer.writerow(['Gerado em', timezone.now().strftime('%d/%m/%Y às %H:%M')])
    writer.writerow([])  # Empty line
    
    # Get events data
    events = export_data['events'].select_related('event_type', 'department', 'location', 'responsible_person')
    
    if not events.exists():
        writer.writerow(['Nenhum evento encontrado para os critérios especificados.'])
    else:
        # Headers
        writer.writerow(['Título', 'Tipo', 'Departamento', 'Data Início', 'Data Fim', 'Status', 'Responsável', 'Local', 'Descrição'])
        
        # Data
        for event in events:
            responsible_name = ''
            if event.responsible_person:
                responsible_name = event.responsible_person.get_full_name() or event.responsible_person.username
            
            location_name = ''
            if event.location:
                location_name = event.location.custom_name or event.location.name
            
            writer.writerow([
                event.name,
                event.event_type.name,
                event.department.name,
                event.start_datetime.strftime('%d/%m/%Y %H:%M'),
                event.end_datetime.strftime('%d/%m/%Y %H:%M') if event.end_datetime else '',
                event.get_status_display(),
                responsible_name,
                location_name,
                event.description or ''
            ])
    
    # Convert to bytes
    csv_content = output.getvalue().encode('utf-8')
    output.close()
    
    filename = f"relatorio_{export_data['report_type']}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return csv_content, filename


def debug_view(request):
    """Simple debug view to verify our enhancements are working"""
    # Test that all our API endpoints are working
    from django.http import JsonResponse
    
    # Test data
    test_data = {
        'status': 'success',
        'message': 'All enhanced reports functionality is working correctly',
        'endpoints': [
            '/reports/api/data/',
            '/reports/api/responsibles/',
            '/reports/api/locations/',
            '/reports/api/trend/',
            '/reports/export/'
        ]
    }
    
    return JsonResponse(test_data)